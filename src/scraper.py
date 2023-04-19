import argparse
from datetime import datetime, timedelta
import logging
import json
import requests
from openpyxl import load_workbook
import xlsxwriter
from fuzzywuzzy import fuzz
import math

logging.basicConfig(level=logging.DEBUG,
format='%(asctime)s %(levelname)s %(message)s',
      filename='./data/cprit.log',
      filemode='w')

AWARD_INFO=['id',
            'agency',
            'awardeeName',
            'startDate',
            'expDate',
            'estimatedTotalAmt',
            'piFirstName',
            'piLastName',
            'pdPIName',
            'title',
            'coPDPI',
            'taccPDPI',
           ]

class Scraper:
    def __init__(self,start,end):
        self.start = start
        self.end = end

    def writeUsers(self,userlist,output,awards):
        """
        Given a list of award information and a list of TACC usernames, write
        an output workbook with two worksheets: (1) Awards that match a TACC username
        and (2) awards that don't match a TACC username.
        """
        userlist_wb = load_workbook(filename=userlist, read_only=True)
        worksheet = userlist_wb['utrc_institution_accounts']
        row_count = worksheet.max_row
        rows = worksheet.rows

        name_dict = {}

        if row_count > 1:
            next(rows) # skip header row
            for row in rows:
                institution = row[0].value
                first_name = row[1].value
                last_name = row[2].value
                name = ' '.join([first_name, last_name]).lower().replace(' ','')
                name_dict[name] = [institution, first_name, last_name]

        logging.info(f'number of items in name_dict = {len(name_dict.keys())}')

        workbook = xlsxwriter.Workbook(output)
        bold = workbook.add_format({'bold': 1})
        found_worksheet = workbook.add_worksheet('utrc_nih_funding')
        found_worksheet.write_row(0, 0, ['utrc_institution', 'utrc_first_name', 'utrc_last_name']+AWARD_INFO, bold)
        not_found_worksheet = workbook.add_worksheet('not_utrc_nih_funding')
        not_found_worksheet.write_row(0, 0, AWARD_INFO, bold)

        f_format = workbook.add_format({'bg_color':'#90EE90'})
        nf_format = workbook.add_format({'bg_color':'#FCC981'})

        f_row = 1
        nf_row = 1
        fuzzy_names = 0
        saved_names = 0

        for item in awards:    
            collab_format = workbook.add_format({'font_color':'red'})

            name_str = item['piFirstName'].lower() + item['piLastName'].lower()
            name_str = name_str.replace(" ", "")
            first_name_str = item['piFirstName'].lower()
            last_name_str = item['piLastName'].lower()
            affiliation = item['awardeeName']

            collaborators = []
            formattedCollab = []
            collab_str = ""

            if(item['coPDPI']!= "NO DATA AVAILABLE"):
                collaborators = item['coPDPI']

            # If a collaborator is in the TACC system, save it for proper formatting.
            # Check for fuzzywuzzy name matching on collaborators.
                
            if(collaborators):
                for z in collaborators:
                    collab_str = z['first_name'].lower() + z['last_name'].lower()
                    if collab_str in name_dict.keys():
                        formattedCollab.append(z['first_name'] + " " + z['last_name'])
                    else:
                        for x in name_dict:
                            if(z['last_name'].lower() != name_dict[x][2].lower()):
                                continue
                            y = fuzz.ratio(z['first_name'].lower(),name_dict[x][1].lower())
                            if(y >= 89 and y < 100 ):
                                fuzzy_names += 1
                                saved_names += 1
                                logging.warning(f"Collaborator {z['first_name']} {z['last_name']} was found based on fuzzywuzzy ratio")
                                formattedCollab.append(name_dict[x][1] + " " + name_dict[x][2])
                                collab_format = workbook.add_format({'bg_color': '#90EE90', 'font_color' : 'red'})


            # If the name matches one in our TACC system, add it to the found sheet. 
            # If the collaborators are in our TACC systems, highlight their names red.

            if name_str in name_dict.keys():
                logging.info(f'{name_str} matches {name_dict[name_str]}')
                found_worksheet.write_row(f_row, 0, [name_dict[name_str][0],
                                                    name_dict[name_str][1],
                                                    name_dict[name_str][2],
                                                    item['id'],
                                                    item['agency'],
                                                    item['awardeeName'],
                                                    item['startDate'],
                                                    item['expDate'],
                                                    item['estimatedTotalAmt'],
                                                    item['piFirstName'],
                                                    item['piLastName'],
                                                    item['pdPIName'],
                                                    item['title'],
                                                    json.dumps(item['coPDPI'])
                                                    ])
                if(formattedCollab):
                    found_worksheet.write(f_row,14,json.dumps(formattedCollab),collab_format)
                else:
                    found_worksheet.write(f_row,14,"None Found")
                f_row += 1

            # If the name does not match one in our TACC system, but a collaborator does, add it to
            # the found sheet. Collaborator will be highlighted in red.

            elif formattedCollab:
                found_worksheet.write_row(f_row, 0, [name_dict[formattedCollab[0].lower().replace(" ","")][0],
                                                    item['piFirstName'],
                                                    item['piLastName'],
                                                    item['id'],
                                                    item['agency'],
                                                    item['awardeeName'],
                                                    item['startDate'],
                                                    item['expDate'],
                                                    item['estimatedTotalAmt'],
                                                    item['piFirstName'],
                                                    item['piLastName'],
                                                    item['pdPIName'],
                                                    item['title'],
                                                    ])
                found_worksheet.write(f_row,13,json.dumps(item['coPDPI']),collab_format)
                found_worksheet.write(f_row,14,json.dumps(formattedCollab),collab_format)
                f_row += 1

            # If the name does not match one in our TACC system, we will search through names that have an exact 
            # last name match. The first name will be compared using fuzzywuzzy word matching. If this returns 
            # a score of 89 or higher, we will pass the PI as a match.

            else:
                logging.info(f'{name_str} has no match')
                
                fuzzy = False
                following = True
                added = False

                for x in name_dict:
                    if(last_name_str != name_dict[x][2].lower()):
                        continue
                    y = fuzz.ratio(first_name_str,name_dict[x][1].lower())
                    if(y >= 80):
                        fuzzy_names += 1
                        fuzzy = True
                        logging.warning(f"Ratio of {y} for {first_name_str} {last_name_str} and {name_dict[x][1].lower()} {name_dict[x][2].lower()}")
                        logging.warning(f"PI Affiliation: {affiliation} && TACC User Affiliation: {name_dict[x][0]}")
                        if(y >= 89 and y < 100 ):
                            saved_names += 1
                            logging.warning(f"Moving {first_name_str} {last_name_str} into sheet (i) based on fuzzywuzzy ratio")
                            if  not added:
                                found_worksheet.write_row(f_row, 0, [name_dict[x][0],
                                                            name_dict[x][1],
                                                            name_dict[x][2],
                                                            item['id'],
                                                            item['agency'],
                                                            item['awardeeName'],
                                                            item['startDate'],
                                                            item['expDate'],
                                                            item['estimatedTotalAmt'],
                                                            item['piFirstName'],
                                                            item['piLastName'],
                                                            item['pdPIName'],
                                                            item['title'],
                                                            json.dumps(item['coPDPI']),
                                                            "None Found"
                                                            ],f_format)
                                f_row += 1
                                following = False
                                added = True
                        
                            
                if following:
                    if fuzzy:
                        format = nf_format
                    else:
                        format = None

                    not_found_worksheet.write_row(nf_row, 0,[ item['id'],
                                                                item['agency'],
                                                                item['awardeeName'],
                                                                item['startDate'],
                                                                item['expDate'],
                                                                item['estimatedTotalAmt'],
                                                                item['piFirstName'],
                                                                item['piLastName'],
                                                                item['pdPIName'],
                                                                item['title'],
                                                                json.dumps(item['coPDPI']),
                                                                "None Found"
                                                            ], format)                                                       
                    nf_row += 1

        found = f_row - 1
        notFound = nf_row - 1  
        if(notFound != 0):
            logging.info("TACC Percentage: {:.2f}".format(float(found/notFound) * 100) + "%")
        logging.info(f'Total fuzzy names: {fuzzy_names}' )
        logging.info(f'Fuzzy names saved: {saved_names}' )

        workbook.close()
        return

    def partition(self,l, n):
        for i in range(0, len(l), n):
            yield l[i:i + n]